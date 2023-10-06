#import package
import openpyxl
#Load Workbook
wk=openpyxl.load_workbook("C:\\Users\\kdruthiv\\PycharmProjects\\RobotFrameworkPrac\\ExcelSheets\\Excel1.xlsx")
sh=wk['Sheet1']
#Read One Cell Data from Excel
print(sh['A3'].value)
print(sh['b2'].value)

#(row,column)
c1=sh.cell(3,2)
c2=sh.cell(column=1,row=4)
print(c1.value)
print(c2.value)
print(c1.row)
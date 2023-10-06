import openpyxl

wk=openpyxl.load_workbook("C:\\Users\\kdruthiv\\PycharmProjects\\RobotFrameworkPrac\\ExcelSheets\\Excel1.xlsx")
def fetch_num_of_rows_sheet(sheetname):
    sh=wk[sheetname]
    return sh.max_row
def fetch_cell_data(sheetname,row,cell):
    sh=wk[sheetname]
    cell=sh.cell(int(row),int(cell))
    x=cell.value
    print(x)
#x=fetch_num_of_rows_sheet("Sheet1")
#print(x)
#print(fetch_cell_data("Sheet1",1,2))
